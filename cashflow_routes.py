"""cashflow_routes.py

FastAPI роуты для модуля подтверждения наличных.

ВАЖНО: модуль специально реализован с "ленивым" импортом функций из app.py,
чтобы не создать циклический импорт при подключении:

    # app.py
    from cashflow_routes import router as cashflow_router
    APP.include_router(cashflow_router)

    import cashflow_models
    # внутри init_db(): cashflow_models.init_cashflow_db(conn)

    import cashflow_bot
    # внутри lifespan после создания bot: cashflow_bot.set_bot(bot)
"""

from __future__ import annotations

import io
import os
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query, Request
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

import cashflow_models as m
import cashflow_bot as b


router = APIRouter()


# ---------------------------
# Dependencies (lazy import from app.py)
# ---------------------------


def _app_db_connect():
    from app import db_connect  # lazy
    return db_connect


def _app_get_current_user():
    from app import get_current_user  # lazy
    return get_current_user


def _app_require_role(*roles: str):
    from app import require_role  # lazy
    return require_role(*roles)


def _get_cfg_paths() -> Dict[str, Any]:
    from app import CFG  # lazy
    return {
        "base_dir": Path(__file__).resolve().parent,
        "users_json": Path(CFG.USERS_JSON_PATH),
        "db_path": Path(CFG.DB_PATH),
    }


def _ensure_cashflow_tables() -> None:
    db_connect = _app_db_connect()
    with db_connect() as conn:
        m.init_cashflow_db(conn)


def _cash_cfg() -> m.CashflowConfig:
    paths = _get_cfg_paths()
    # переопределяем paths, чтобы совпадало с app.py
    cfg = m.load_cashflow_config(paths["base_dir"])
    return m.CashflowConfig(
        base_dir=cfg.base_dir,
        db_path=paths["db_path"],
        users_json_path=paths["users_json"],
        uploads_dir=cfg.uploads_dir,
        timezone=cfg.timezone,
    )


# ---------------------------
# Models
# ---------------------------


class CashRequestCreateIn(BaseModel):
    account: str = Field(..., description="main|praise|alpha")
    op_type: str = Field(..., description="collect|withdraw")
    amount: float
    source_kind: Optional[str] = None
    source_id: Optional[int] = None


class CashResendIn(BaseModel):
    target_telegram_ids: Optional[List[int]] = Field(None, description="Если не задано — всем отказавшим на попытке 1")
    admin_comment: Optional[str] = None


class CashSignIn(BaseModel):
    signature: str = Field(..., description="data:image/png;base64,... или base64")


class CashRefuseIn(BaseModel):
    reason: str = Field(..., min_length=1)


# ---------------------------
# HTML entry (cashapp)
# ---------------------------


@router.get("/cashapp")
def cashapp_html():
    """Отдельный интерфейс подписанта (не бухгалтерия)."""
    path = Path(__file__).resolve().parent / "cashapp.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="cashapp.html not found")
    return FileResponse(str(path), media_type="text/html")


# ---------------------------
# API: Requests
# ---------------------------


@router.post("/api/cashflow/requests")
def create_request(
    body: CashRequestCreateIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("admin", "accountant")),
):
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()

    with db_connect() as conn:
        request_id = m.create_cash_request(
            conn,
            cfg,
            account=body.account,
            op_type=body.op_type,
            amount=float(body.amount),
            created_by_telegram_id=int(u["telegram_id"]),
            source_kind=body.source_kind,
            source_id=body.source_id,
        )
        view = m.build_request_view(conn, request_id)
        req = view["request"]
        signers = [p["telegram_id"] for p in view["participants"] if not p["is_admin"]]

    # уведомляем подписантов
    bg.add_task(
        b.notify_signers_new_request,
        request_id=request_id,
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_ids=signers,
        is_retry=False,
        admin_comment=None,
    )
    return {"id": request_id, "item": view}


@router.get("/api/cashflow/requests/my")
def my_requests(
    account: Optional[str] = Query(None),
    only_open: bool = Query(False),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        items = m.list_my_cash_requests(
            conn,
            telegram_id=int(u["telegram_id"]),
            account=account,
            only_open=bool(only_open),
            limit=limit,
            offset=offset,
        )
        return {"items": [dict(x) for x in items]}


@router.get("/api/cashflow/requests")
def list_requests(
    account: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        items = m.list_cash_requests(conn, account=account, status=status, limit=limit, offset=offset)
        return {"items": [dict(x) for x in items]}


@router.get("/api/cashflow/requests/{request_id}")
def request_detail(
    request_id: int,
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        view = m.build_request_view(conn, int(request_id))
        # доступ: админ или участник
        if str(u["role"]) != "admin":
            tid = int(u["telegram_id"])
            if not any(int(p["telegram_id"]) == tid for p in view["participants"]):
                raise HTTPException(status_code=403, detail="Not a participant")
        return view


# ---------------------------
# API: Sign / Refuse
# ---------------------------


@router.post("/api/cashflow/requests/{request_id}/sign")
def sign_request(
    request_id: int,
    body: CashSignIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("cash_signer", "admin")),
):
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        # admin должен подписывать через /admin-sign
        if str(u["role"]) == "admin":
            raise HTTPException(status_code=403, detail="Admin must use /admin-sign")
        try:
            m.record_signature(
                conn,
                cfg,
                request_id=int(request_id),
                telegram_id=int(u["telegram_id"]),
                signature_data_url=body.signature,
                as_admin=False,
            )
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

        view = m.build_request_view(conn, int(request_id))
        req = view["request"]
        # найдём имя подписанта
        signer_name = str(u["name"])

    # уведомить ответственного админа
    bg.add_task(
        b.notify_admin_about_decision,
        admin_id=int(req["admin_telegram_id"]),
        request_id=int(request_id),
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_name=signer_name,
        decision="SIGNED",
        reason=None,
    )
    # если финализировано — уведомить инициатора
    if req.get("created_by_telegram_id") and req.get("status") == "FINAL":
        bg.add_task(
            b.notify_initiator_final,
            initiator_id=int(req["created_by_telegram_id"]),
            request_id=int(request_id),
            account=req["account"],
            op_type=req["op_type"],
            amount=float(req["amount"]),
            status=str(req["status"]),
        )

    return {"ok": True, "item": view}


@router.post("/api/cashflow/requests/{request_id}/refuse")
def refuse_request(
    request_id: int,
    body: CashRefuseIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("cash_signer")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.record_refusal(
                conn,
                request_id=int(request_id),
                telegram_id=int(u["telegram_id"]),
                reason=str(body.reason).strip(),
            )
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

        view = m.build_request_view(conn, int(request_id))
        req = view["request"]
        signer_name = str(u["name"])

    bg.add_task(
        b.notify_admin_about_decision,
        admin_id=int(req["admin_telegram_id"]),
        request_id=int(request_id),
        account=req["account"],
        op_type=req["op_type"],
        amount=float(req["amount"]),
        signer_name=signer_name,
        decision="REFUSED",
        reason=str(body.reason).strip(),
    )
    return {"ok": True, "item": view}


# ---------------------------
# API: Admin actions
# ---------------------------


@router.post("/api/cashflow/requests/{request_id}/resend")
def admin_resend(
    request_id: int,
    body: CashResendIn = Body(default_factory=CashResendIn),
    bg: BackgroundTasks = None,  # type: ignore[assignment]
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    cfg = _cash_cfg()
    with db_connect() as conn:
        try:
            targets = m.resend_for_refusals(
                conn,
                request_id=int(request_id),
                admin_telegram_id=int(u["telegram_id"]),
                target_telegram_ids=body.target_telegram_ids,
                admin_comment=body.admin_comment,
            )
            view = m.build_request_view(conn, int(request_id))
            req = view["request"]
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    if bg is not None:
        bg.add_task(
            b.notify_signers_new_request,
            request_id=int(request_id),
            account=req["account"],
            op_type=req["op_type"],
            amount=float(req["amount"]),
            signer_ids=targets,
            is_retry=True,
            admin_comment=req.get("admin_comment"),
        )
    return {"ok": True, "targets": targets, "item": view}


@router.post("/api/cashflow/requests/{request_id}/admin-sign")
def admin_sign(
    request_id: int,
    body: CashSignIn,
    bg: BackgroundTasks,
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.record_signature(
                conn,
                cfg,
                request_id=int(request_id),
                telegram_id=int(u["telegram_id"]),
                signature_data_url=body.signature,
                as_admin=True,
            )
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

        view = m.build_request_view(conn, int(request_id))
        req = view["request"]

    if req.get("created_by_telegram_id") and req.get("status") == "FINAL":
        bg.add_task(
            b.notify_initiator_final,
            initiator_id=int(req["created_by_telegram_id"]),
            request_id=int(request_id),
            account=req["account"],
            op_type=req["op_type"],
            amount=float(req["amount"]),
            status=str(req["status"]),
        )
    return {"ok": True, "item": view}


@router.post("/api/cashflow/requests/{request_id}/cancel")
def admin_cancel(
    request_id: int,
    comment: Optional[str] = Body(None),
    u=Depends(_app_require_role("admin")),
):
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        try:
            m.cancel_request(conn, request_id=int(request_id), admin_telegram_id=int(u["telegram_id"]), comment=comment)
            view = m.build_request_view(conn, int(request_id))
        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, "item": view}


# ---------------------------
# API: Signature image
# ---------------------------


@router.get("/api/cashflow/requests/{request_id}/participants/{telegram_id}/signature.png")
def get_signature_png(
    request_id: int,
    telegram_id: int,
    u=Depends(_app_require_role("admin", "accountant", "viewer", "cash_signer")),
):
    """Возвращает эффективную подпись участника: attempt=2 если есть, иначе attempt=1."""
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        view = m.build_request_view(conn, int(request_id))
        # доступ: админ или участник
        if str(u["role"]) != "admin":
            tid = int(u["telegram_id"])
            if not any(int(p["telegram_id"]) == tid for p in view["participants"]):
                raise HTTPException(status_code=403, detail="Not a participant")

        target = None
        for p in view["participants"]:
            if int(p["telegram_id"]) == int(telegram_id):
                target = p
                break
        if not target:
            raise HTTPException(status_code=404, detail="Participant not found")

        sig2 = target.get("attempt2")
        sig1 = target.get("attempt1")
        sig = sig2 or sig1
        if not sig or sig.get("decision") != "SIGNED" or not sig.get("signature_path"):
            raise HTTPException(status_code=404, detail="Signature not found")
        path = m.get_signature_file_path(cfg, str(sig["signature_path"]))
        if not path.exists():
            raise HTTPException(status_code=404, detail="Signature file missing")
        return FileResponse(str(path), media_type="image/png")


# ---------------------------
# API: Withdrawal Act (view + Excel)
# ---------------------------


def _user_can_view_withdraw_act(telegram_id: int, role: str) -> bool:
    if role == "admin":
        return True
    cfg = _cash_cfg()
    allow = m.load_users_allowlist(cfg.users_json_path)
    u = allow.get(int(telegram_id))
    if not u or not (u.get("active") is True or u.get("active") == 1):
        return False
    ops = u.get("cash_ops") or []
    if isinstance(ops, str):
        ops = [x.strip() for x in ops.split(",") if x.strip()]
    return str(u.get("role")) == "cash_signer" and "withdraw" in [str(x).lower() for x in ops]


@router.get("/api/cashflow/withdraw-act")
def withdraw_act(
    account: Optional[str] = Query(None, description="main|praise|alpha (опционально)"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    u=Depends(_app_require_role("admin", "cash_signer")),
):
    if not _user_can_view_withdraw_act(int(u["telegram_id"]), str(u["role"])):
        raise HTTPException(status_code=403, detail="No access to withdraw act")
    _ensure_cashflow_tables()
    db_connect = _app_db_connect()
    with db_connect() as conn:
        rows = m.list_withdraw_act_rows(conn, account=account, date_from=date_from, date_to=date_to)
    return {"items": rows}


@router.get("/api/cashflow/withdraw-act.xlsx")
def withdraw_act_xlsx(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    u=Depends(_app_require_role("admin")),
):
    """Экспорт акта изъятия в Excel.

    Формат: одна книга, 3 листа (MAIN/PRAISE/ALPHA), на каждом 5 колонок:
    Дата, Сумма Изъятия, ФИО, Тип пользователя, Подпись.
    """
    _ensure_cashflow_tables()
    cfg = _cash_cfg()
    db_connect = _app_db_connect()

    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise HTTPException(status_code=501, detail=f"openpyxl is not available: {e}")

    with db_connect() as conn:
        rows_all = m.list_withdraw_act_rows(conn, account=None, date_from=date_from, date_to=date_to)

    wb = openpyxl.Workbook()
    # удаляем дефолтный лист
    wb.remove(wb.active)

    def make_sheet(account_code: str, title: str):
        ws = wb.create_sheet(title)
        ws.append(["Дата", "Сумма Изъятия", "ФИО", "Тип пользователя", "Подпись"])
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 32

        rnum = 2
        for r in rows_all:
            if r["account"] != account_code:
                continue
            ws.cell(row=rnum, column=1, value=r.get("date"))
            ws.cell(row=rnum, column=2, value=float(r.get("amount") or 0))
            ws.cell(row=rnum, column=3, value=r.get("fio"))
            ws.cell(row=rnum, column=4, value=r.get("user_type"))

            sig_cell = ws.cell(row=rnum, column=5)
            if r.get("signature_value", "").startswith("ОТКАЗ"):
                sig_cell.value = r.get("signature_value")
            elif r.get("signature_path"):
                # вставляем картинку подписи
                try:
                    img_path = m.get_signature_file_path(cfg, str(r["signature_path"]))
                    if img_path.exists():
                        img = XLImage(str(img_path))
                        # приводим к разумному размеру
                        img.width = 240
                        img.height = 70
                        anchor = f"{get_column_letter(5)}{rnum}"
                        ws.add_image(img, anchor)
                        ws.row_dimensions[rnum].height = 55
                    else:
                        sig_cell.value = "SIGNED"
                except Exception:
                    sig_cell.value = "SIGNED"
            else:
                sig_cell.value = "SIGNED"
            rnum += 1

    make_sheet("main", "MAIN")
    make_sheet("praise", "PRAISE")
    make_sheet("alpha", "ALPHA")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "withdraw_act.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
